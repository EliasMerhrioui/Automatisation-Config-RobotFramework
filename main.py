import os
import openpyxl
import json

# Lecture des informations depuis le fichier Excel
config_wb = openpyxl.load_workbook('./Configuration/config_API.xlsx')
config_sheet = config_wb.active

api_name = config_sheet['A2'].value
api_version = config_sheet['A4'].value
api_env = config_sheet['A6'].value


#-------------------ICI AJOUTER LES d01,d02...--------------------------
# Création de l'URL en fonction de la version et de l'environnement
if api_env == "dev":
    url = f"url ici"
elif api_env == "test":
    url = f"url ici"
elif api_env == "essai":
    url = f"url ici"
else:
    print("Environnement invalide")

# Extraction des noms d'endpoints
endpoints_str = config_sheet['A8'].value
endpoints_list = endpoints_str.split(',')
nb_endpoints = len(endpoints_list)


# Mise à jour des noms d'endpoints et des cellules adjacentes vides
for i, endpoint_name in enumerate(endpoints_list):
    config_sheet.cell(row=9+i*2, column=1).value = f"Endpoint numéro {i+1} : {endpoint_name}"
    config_sheet.cell(row=9+i*2, column=2).value = "json"

# Sauvegarde des modifications dans le fichier Excel
config_wb.save('./Configuration/config_API.xlsx')

# Création du dossier pour l'API
os.makedirs(api_name, exist_ok=True)

# Création des sous-dossiers et fichiers
os.makedirs(os.path.join(api_name, "Bibliothèques"), exist_ok=True)
os.makedirs(os.path.join(api_name, "Config"), exist_ok=True)
os.makedirs(os.path.join(api_name, "Fichiers"), exist_ok=True)
os.makedirs(os.path.join(api_name, "Scenarios"), exist_ok=True)

open(os.path.join(api_name, "interactive_console_output.xml"), "a").close()
open(os.path.join(api_name, "log.html"), "a").close()
open(os.path.join(api_name, "output.xml"), "a").close()
open(os.path.join(api_name, "report.xml"), "a").close()

input("Tous les paramètres des endpoints sont ok ? Appuyez sur Entrée pour continuer...")

# Créer un dictionnaire pour stocker les données de chaque endpoint
endpoint_data = {}

with open(os.path.join(api_name, "Bibliothèques", "fonctionnel.resource"), "a") as fr:
    fr.write('*** Settings ***\nResource          MGEN-Keywords/API.resource\nLibrary           Collections\nLibrary    String\n\n*** Keywords ***\n')

# Pour chaque nom d'endpoint dans la liste endpoints_list
for i, endpoint_name in enumerate(endpoints_list):
    # Trouver la ligne qui contient le nom de l'endpoint
    for row in config_sheet.iter_rows(min_row=9+i*2):
        if row[0].value == f"Endpoint numéro {i+1} : {endpoint_name}":
            # Extraire les données situées sous la ligne de l'endpoint
            data = []
            for row_data in config_sheet.iter_rows(min_row=row[0].row+1):
                if row_data[0].value is None:
                    break
                if i+1 < len(endpoints_list) and row_data[0].value != f"Endpoint numéro {i+2} : {endpoints_list[i+1]}":
                    if not row_data[0].value.startswith('Endpoint numéro'):
                        data.append(row_data[0].value)
                else:
                    if not row_data[0].value.startswith('Endpoint numéro'):
                        data.append(row_data[0].value)
                    break
    
            # Enregistrer les données extraites dans le dictionnaire endpoint_data
            endpoint_data[endpoint_name] = data

            open(os.path.join(api_name, "Scenarios", f"{endpoint_name}.robot"), "a").close()
            open(os.path.join(api_name, "Config", f"{endpoint_name}.yaml"), "w")            
           
            # Création du fichier fonctionnel.resource
            with open(os.path.join(api_name, "Bibliothèques", "fonctionnel.resource"), "a+") as fr:
                fr.write(endpoint_name + '\n')
                fr.write('    [Arguments]    ${method}    ${nom du test}\n')
                fr.write('    ${Body}    create Dictionary\n')
                fr.write('    Set test variable    ${Body}\n')
                
                params_iter = iter(endpoint_data[endpoint_name])
                while True:
                    try:
                        param = next(params_iter)
                        params_list = param.split(',')
                        for param in params_list:
                            fr.write(f'    Ajout    {param}    ${{{param}}}\n')
                    except StopIteration:
                        break 
                    fr.write('    ${response}    Run keyword    ${method}    ${body}    ${nom du test}')
                    fr.write('\n')
                    # fr.write(r'    Set Tags    code retour ${response.status_code} \r\n Fichier JSON :\r\n ${response.json()}')
                    # fr.write('\n')
                    # fr.write(r'    Set Test documentation    code retour ${response.status_code} \r\n Fichier JSON :\r\n ${response.json()}')
                    fr.write('\n')   
    
    with open(os.path.join(api_name, "Bibliothèques", "fonctionnel.resource"), "r+") as fr:
        lines = fr.readlines()
        endpoint_line_index = -1
        for i, line in enumerate(lines):
            if line.strip() == "[Arguments]    ${method}    ${nom du test}":
                endpoint_line_index = i
                break

        if endpoint_line_index != -1:
            # Récupération des paramètres du endpoint
            params_var = []
            for param in endpoint_data[endpoint_name]:
                for p in param.split(','):
                    params_var.append('${' + p.strip() + '}')
            if len(params_var) > 0:
                lines[endpoint_line_index] = lines[endpoint_line_index].rstrip() + "    " + "    ".join(params_var) + "\n"

            # Écriture des modifications dans le fichier fonctionnel.resource
            fr.seek(0)
            fr.writelines(lines)
    
    #Configuration des fichiers .yaml
    filename = os.path.join(api_name, "Config", endpoint_name + ".yaml")
    with open(filename, "a") as f:
        f.write("url: {}\n".format(url))
        f.write("api: {}-front/{}/{}\n".format(api_name, api_version, endpoint_name))
        f.write("headers:\n")        
        f.write("  Content-Type: application/json\n")
        f.write("  accept: application/json\n")
        f.write("TokenUrl: urlici/\n") #optimisation: SI c'est dev test essai
        f.write("TokenEndpoint: cjwt-bff_utils/create-token-jwtgen/\n")
        f.write("TokenEnv: DEV\n") #optimisation: SI c'est dev test essai


# Dictionnaire pour stocker les propriétés de chaque endpoint
endpoint_properties = {}


# Récupérer l'incrément souhaité par l'utilisateur depuis le fichier Excel
increment = config_sheet['B2'].value

# Fonction récursive pour générer les noms de propriétés des tableaux
def generate_property_names(data, prefix='', increment=1):
    properties = []
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, dict):
                properties.extend(generate_property_names(value, prefix + key + '_', increment))
            elif isinstance(value, list):
                if value:
                    item = value[0]
                    if isinstance(item, dict):
                        properties.extend(generate_property_names(item, prefix + key + '_0_', increment))
                        for i in range(1, increment):
                            properties.extend(generate_property_names(item, prefix + key + '_{}_'.format(i), increment))
                    else:
                        if isinstance(value[0], dict):
                            properties.extend(generate_property_names(value[0], prefix + key + '_0_', increment))
            else:
                properties.append(prefix + key)
    elif isinstance(data, list):
        if data:
            item = data[0]
            if isinstance(item, dict):
                properties.extend(generate_property_names(item, prefix + '{}_'.format(0), increment))
    return properties

# Parcourir les noms d'endpoints
for endpoint_name in endpoints_list:
    json_doc = {}
    for row in config_sheet.iter_rows():
        if row[1].value == "json":
            endpoint_name_excel = row[0].value.split(":")[1].strip()
            data = []
            for row_data in config_sheet.iter_rows(min_row=row[1].row + 1):
                if row_data[0].value is None:
                    break
                data.append(row_data[1].value)
            json_doc[endpoint_name_excel] = data[0]

    # Charger le JSON en tant que dictionnaire Python
    data = json.loads(json_doc[endpoint_name])

    # Fonction pour vérifier si deux dictionnaires sont identiques
    def are_dicts_equal(dict1, dict2):
        if set(dict1.keys()) != set(dict2.keys()):
            return False
        for key in dict1.keys():
            if isinstance(dict1[key], dict):
                if not are_dicts_equal(dict1[key], dict2[key]):
                    return False
            elif dict1[key] != dict2[key]:
                return False
        return True

    # Extraire les propriétés des tableaux dans le dictionnaire
    properties = generate_property_names(data, increment=increment)

    # Filtrer les propriétés pour ne garder que celles des premiers documents
    filtered_properties = []
    for property_name in properties:
        property_parts = property_name.split('_')
        if property_parts[-1].isdigit():
            index = int(property_parts[-1])
            if index == 0:
                filtered_properties.append(property_name)
            else:
                prefix = '_'.join(property_parts[:-1])
                previous_property = '_'.join([prefix, str(index - 1)])
                if not are_dicts_equal(data, endpoint_properties[endpoint_name]):
                    filtered_properties.append(property_name)
        else:
            filtered_properties.append(property_name)
    # Stocker les propriétés de chaque endpoint dans un dictionnaire
    endpoint_properties[endpoint_name] = filtered_properties

# Créer un dictionnaire pour sauvegarder les propriétés par endpoint
properties_by_endpoint = {}
# print(endpoint_properties)

# ------------------------


# Chemin du dossier Scenarios
dossier_scenarios = os.path.join(api_name, "Scenarios")
# Parcourir la liste des noms d'endpoint
for endpoint_name in endpoints_list:
    # Construire le chemin complet du fichier .robot
    chemin_fichier = os.path.join(dossier_scenarios, f"{endpoint_name}.robot")
    
    with open(chemin_fichier, "a", encoding="utf-8") as fichier:
        fichier.write("*** Settings *** \n")
        fichier.write("Test Template     Cas Passants \n")
        fichier.write("Library           DataDriver    ../Fichiers/CasDeTest_API_{}.xls    sheet_name={} \n".format(api_name, endpoint_name))
        fichier.write("Resource          ../Bibliothèques/fonctionnel.resource \n")
        fichier.write("Variables         ../Config/{}.yaml \n".format(endpoint_name))
        fichier.write("\n")
        fichier.write("*** Test Cases *** \n")
        fichier.write("${nom du test}\n")
        fichier.write("\n")
        fichier.write("*** Keywords *** \n")
        fichier.write("Cas Passants\n")
        fichier.write("    [Arguments]    ${nom du test}\n")
    
# Parcourir la liste des noms d'endpoint 
for endpoint_name in endpoints_list:
    # Construire le chemin complet du fichier .robot
    chemin_fichier = os.path.join(dossier_scenarios, f"{endpoint_name}.robot")

    # Ouvrir le fichier en mode lecture/écriture
    with open(chemin_fichier, "r+", encoding="utf-8") as fichier:
        # Lire le contenu du fichier
        contenu_fichier = fichier.readlines()

        # Rechercher la ligne contenant "Cas Passants" dans le contenu du fichier
        for i, ligne in enumerate(contenu_fichier):
            if "    [Arguments]    ${nom du test}" in ligne:
                # Extraire les paramètres de endpoint_data[endpoint_name]
                params = endpoint_data[endpoint_name]

                # Ajouter les éléments du paramètre un par un en utilisant la boucle for
                for param in params:
                    param_elements = param.split(',')
                    contenu_fichier[i] = ligne.rstrip() + "    " + "    ".join([f"${{{param_element.strip()}}}" for param_element in param_elements]) + "\n"
                break

        # Se positionner au début du fichier pour réécrire le contenu
        fichier.seek(0)
        # Écrire le contenu modifié dans le fichier
        fichier.writelines(contenu_fichier)
        # Tronquer le reste du fichier au niveau de la dernière position d'écriture
        fichier.truncate()

        # Afficher un message pour indiquer que la modification a été effectuée
        # print(f"Contenu ajouté dans le fichier : {chemin_fichier}")

# Ajout de tout les endpoint_properties a la suite 
for endpoint_name in endpoints_list:
    # Construire le chemin complet du fichier .robot
    chemin_fichier = os.path.join(dossier_scenarios, f"{endpoint_name}.robot")

    # Ouvrir le fichier en mode append
    with open(chemin_fichier, "a", encoding="utf-8") as fichier:
        # Extraire les paramètres de endpoint_properties[endpoint_name]
        params = endpoint_properties[endpoint_name]

        # Vérifier si le nombre de paramètres dépasse 8
        if len(params) > 8:
            # Écrire la première partie des paramètres sur la même ligne
            fichier.write("    " + "..." + "    " + "    ".join([f"${{{param_element.strip()}}}" for param_element in params[:4]]))
            # Faire un retour à la ligne avec "    " + "..." + "    "
            fichier.write("\n    " + "..." + "    ")
            # Écrire la suite des paramètres avec un maximum de 4 paramètres par ligne
            for i in range(4, len(params), 4):
                fichier.write("    ".join([f"${{{param_element.strip()}}}" for param_element in params[i:i+4]]))
                fichier.write("\n    " + "..." + "    ")

            # Supprimer le "..." à la fin du fichier
            fichier.seek(0, os.SEEK_END)
            fichier.seek(fichier.tell() - len("\n    " + "..." + "    "), os.SEEK_SET)
            fichier.truncate()

            fichier.write("\n")
        else:
            # Écrire les paramètres sur la même ligne
            fichier.write("    " + "..." + "    " + "    ".join([f"${{{param_element.strip()}}}" for param_element in params]))
            fichier.write("\n")

    # Afficher un message pour indiquer que la modification a été effectuée
    #print(f"Contenu ajouté dans le fichier : {chemin_fichier}")

# ---------------------------------------------------------- #


# Ajout du Vérifier temps get etc
for endpoint_name in endpoints_list:
    # Construire le chemin complet du fichier .robot
    chemin_fichier = os.path.join(dossier_scenarios, f"{endpoint_name}.robot")

    # Ouvrir le fichier en mode append
    with open(chemin_fichier, "a", encoding="utf-8") as fichier:
        # Ajouter la ligne avec le nom de l'endpoint et le nom du test
        fichier.write("    " + endpoint_name + "    Get    ${nom du test}" + "\n")
        # Ajouter les nouvelles lignes
        fichier.write("    Vérifier le Temps    300\n")
        fichier.write("    Vérifier le code retour    200\n")


    # Lire le contenu du fichier
    with open(chemin_fichier, "r", encoding="utf-8") as fichier:
        contenu_fichier = fichier.readlines()

    # Parcourir le contenu du fichier
    for i, ligne in enumerate(contenu_fichier):
        if "Get    ${nom du test}" in ligne:
            # Extraire les paramètres de endpoint_data[endpoint_name]
            params = endpoint_data[endpoint_name]

            # Ajouter les éléments du paramètre un par un en utilisant la boucle for
            for param in params:
                param_elements = param.split(',')
                ligne_param = ligne.rstrip() + "    " + "    ".join([f"${{{param_element.strip()}}}" for param_element in param_elements]) + "\n"
                contenu_fichier.insert(i + 1, ligne_param)  # Insérer la ligne avec les param_elements après la ligne actuelle
            contenu_fichier.pop(i)  # Supprimer la ligne originale
            break

    # Écrire le contenu modifié dans le fichier
    with open(chemin_fichier, "w", encoding="utf-8") as fichier:
        fichier.writelines(contenu_fichier)

    # Afficher un message pour indiquer que la modification a été effectuée
    #print("Contenu ajouté dans le fichier : " + chemin_fichier)



# ---------------------------------------------------------- #
# Définition de endpoint_data_list
endpoint_data_list = endpoint_data


# Créer un dossier pour les fichiers s'il n'existe pas
dossier_fichiers = f"./{api_name}/Fichiers"
if not os.path.exists(dossier_fichiers):
    os.makedirs(dossier_fichiers)

# Créer un nouveau fichier Excel
classeur = openpyxl.Workbook()
# Parcourir la liste des noms d'endpoint
for i, endpoint_name in enumerate(endpoints_list):
    # Créer une nouvelle feuille de calcul pour l'endpoint
    feuille_calcul = classeur.create_sheet(title=endpoint_name)
    
    # Écrire les en-têtes des colonnes
    feuille_calcul.cell(row=1, column=1, value="${nom du test}")
    # Changer la couleur de la police en jaune
    feuille_calcul.cell(row=1, column=1).fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Initialiser la liste endpoint_data
    endpoint_data = []

    # Extraire les données de endpoint_data_list[endpoint_name]
    endpoint_data += endpoint_data_list[endpoint_name]

    # Extraire les données de endpoint_properties[endpoint_name]
    properties_data = endpoint_properties[endpoint_name]

    # Ajouter les données de endpoint_properties[endpoint_name] à endpoint_data en les séparant par des virgules
    if properties_data:
        if not endpoint_data[0].endswith(','): # Vérifier si la dernière valeur de endpoint_data[0] est déjà une virgule
            endpoint_data[0] += ',' # Si non, ajouter une virgule
        endpoint_data[0] += ','.join(properties_data) # Ajouter les données de propriétés

    # Parcourir les données de l'endpoint
    for row, data in enumerate(endpoint_data, start=2):
        # Vérifier si les données sont une chaîne de caractères
        if isinstance(data, str):
            # Remplacer les deux-points par des virgules
            data = data.replace(':', ',')
            
            # Diviser les données en clés et valeurs en utilisant la virgule comme séparateur
            items = data.split(',')
            items += endpoint_properties[endpoint_name]
            # print("Data: " + data)
            # print("Items: " + str(items))
            print(items)
            
            # Réinitialiser la variable col pour éviter la duplication des colonnes
            col = 2
            
            # Ajouter les clés (nom du test) dans colonnes correspondantes
            for key in items[0].split(','):
                # Déterminer la colonne dans laquelle écrire les données
                col_letter = openpyxl.utils.get_column_letter(col)
                feuille_calcul.cell(row=1, column=col, value=f'${{{key}}}')
                col += 1
                
            # Parcourir les valeurs dans la liste Items et les ajouter dans les colonnes correspondantes
            for value in items[1:]:
                # Déterminer la colonne dans laquelle écrire les données
                col_letter = openpyxl.utils.get_column_letter(col)
                feuille_calcul.cell(row=row-1, column=col, value=f'${{{value}}}')
                col += 1
            
            # Créer le répertoire Scenarios s'il n'existe pas
            scenarios_folder = "Scenarios"

            # Afficher les items dans le terminal
            print(f"\nItems pour l'endpoint {endpoint_name}:\n")
            for item in items:
                print(item)

                # Ajouter l'item à la fin du fichier .robot
                robot_file_path = os.path.join(dossier_scenarios, f"{endpoint_name}.robot")
                with open(robot_file_path, "a", encoding='utf-8') as fichier_robot:
                    fichier_robot.write(f"{'    Vérifier la valeur  ' if not any(char.isdigit() for char in item) else '    Vérifier la collection    '} {item}     ${{{item}}}\n")




# Enregistrer le fichier Excel
nom_fichier_excel = f"CasDeTest_API_{api_name}.xls"
chemin_fichier_excel = os.path.join(dossier_fichiers, nom_fichier_excel)
classeur.save(chemin_fichier_excel)




























